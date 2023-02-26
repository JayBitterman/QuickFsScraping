from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import time
import openpyxl as xl
from openpyxl import Workbook as WB
from selenium.webdriver.chrome.options import Options

# put any number of ticker symbols in list
# e.g stocks = ['aapl', 'amzn']
stocks = []

# set up driver. You will need to have downloaded the correct chromedriver version
ser = Service("<Path to Chromedriver>")
options = Options()

# Special folder holding my user data on google websites. Allows the driver to 'remember' me.
# Otherwise you need to manually input these credentials into quickfs so log in.
options.add_argument("<Path to MyUser data folder>")

options.page_load_strategy = 'normal'
driver = webdriver.Chrome(service=ser, options=options)

# set up new excel files for each stock
for st_num, stock in enumerate(stocks):
    # use any valid ticker
    ticker = stocks[st_num]
    # this will create the excel file in your documents
    path_cvs = rf"<PathToStoreFile>\{ticker}Val.xlsx"
    wb = WB(path_cvs)
    wb.save(path_cvs)
    wb = xl.load_workbook(path_cvs)
    wb.remove(wb.active)

    s = wb.create_sheet('DATA INPUT')

    # load quickfs.net webste
    driver.get('https://quickfs.net/')

    time.sleep(1)

    # search for ticker symbol
    search = driver.find_element(By.CLASS_NAME, 'navbar-form').find_element(By.TAG_NAME, 'input')
    search.send_keys(ticker)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'acResultItem'))).click()

    # simple helper functions

    def years():
        year_list = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'thead'))).find_elements(By.TAG_NAME, 'td')
        return year_list


    # helps clean data from non-numerical symbols
    def con(str):
        str = str.replace(',', '')
        str = str.replace('$', '')
        str = str.replace('%', '')
        str = str.replace('£', '')
        return float(str)


    # opens dropdown menu
    def c_drop():
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "select-fs-dropdown"))).find_element(By.TAG_NAME, "button").click()


    # sets scale on website to 1000
    def thous():
        drop = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "select-rounding-dropdown")))
        drop.click()
        drop.find_element(By.ID, 'thousands').click()


    # select data fields for each year listed on the website
    for x, year in enumerate(years()):
        s.cell(1, x + 1).value = year.text

    # list of data fields we want and its destination row in the excel sheet for each of the 5 pages respectively
    memo = [
        [["price-to-earnings"], [5]],
        [["eps (diluted)", "shares (diluted)", "revenue", "pre-tax income", "income tax", "operating profit"], [7, 8, 6, 10, 11, 9]],
        [["shareholders' equity", "total current assets", "total current liabilities", "total liabilities", "accounts receivable",
          "accounts payable", "cash & equivalents", "short-term debt", "long-term debt"], [20, 12, 13, 14, 15, 16, 17, 18, 19]],
        [["cash from operations", "property, plant, & equipment", "net income", "depreciation & amortization", "cash paid for dividends"],
         [23, 25, 21, 22, 24]]
    ]

    # the 5 financial statement sheets we will be scrapping
    for page_num, page in enumerate(["ratios", "is", "bs", "cf"]):
        c_drop()
        driver.find_element(By.ID, page).click()
        trs = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'fs-table'))).find_elements(By.TAG_NAME, 'tr')
        for tr in trs:
            try:
                for r, item in enumerate(memo[page_num][0]):
                    field = tr.find_element(By.CLASS_NAME, 'labelCell').text.lower()
                    if item in field:
                        s.cell(memo[page_num][1][r], 1).value = field
                        data = tr.find_elements(By.CLASS_NAME, 'dataCell')
                        if page == 'bs':
                            data.pop(0)
                        for n, datum in enumerate(data):
                            if datum.text == '-':
                                continue
                            cell = s.cell(memo[page_num][1][r], n + 2)
                            cell.value = con(datum.text)
            except NoSuchElementException:
                pass


    # Now get some info from yahoo finance
    # Destroy the adds
    driver.get('https://finance.yahoo.com/')
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "myLightboxContainer"))).find_element(By.TAG_NAME, "path").click()
    except NoSuchElementException:
        pass
    search2 = driver.find_element(By.ID, 'yfin-usr-qry')
    search2.send_keys(ticker)
    search2.submit()
    time.sleep(1)
    try:
        driver.find_element(By.XPATH, '//span[text()="Analysis"]').click()
        # Finding analyst growth on Yahoo Finance
        growth = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//span[text()="Next 5 Years (per annum)"]/..//following-sibling::td'))).text
        s.cell(30, 1).value = "Analyst growth for next 5 years per annum"
        s.cell(30, 2).value = growth
    except NoSuchElementException:
        pass

    wb.save(path_cvs)
    wb.close()
driver.quit()
